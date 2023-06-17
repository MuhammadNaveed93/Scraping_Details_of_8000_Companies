# -*- coding: utf-8 -*-
"""
Created on Wed Jan  4 10:48:21 2023

@author: muhammad.naveed
"""

def Convert_to_Excel(df,filename):

    import os
    import openpyxl
    import pandas as pd
    from openpyxl.utils.dataframe import dataframe_to_rows
    file=r"D:\\Naveed\\Freelancing\\1. Upwork\\2. Work for Portfolio\\2. Jobs to be performed\\Job02\\Program\\Output\\"+filename+".xlsx"
    if os.path.isfile(file):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook["Sheet1"]
        for row in dataframe_to_rows(df,header=False,index=False):
            sheet.append(row)
        workbook.save(file)
        workbook.close()
    else:
        with pd.ExcelWriter(path=file, engine='openpyxl') as writer:
            df.to_excel(writer,index=False,sheet_name="Sheet1")
            
            
def Data_to_Excel(dictionaries_list,filename):
    
    import pandas as pd
    
    data = []
        
    for dictionary in dictionaries_list:
        max_length = max(len(dictionary['personal_data']), len(dictionary['functions']),len(dictionary['signature']))
        other_values = [dictionary['Type'], dictionary['UID'], dictionary['Business_name'], dictionary['Company_address'],dictionary['District']]

        for i in range(max_length):
            row_data = other_values + [dictionary['personal_data'][i], dictionary['functions'][i], dictionary['signature'][i]]
            data.append(row_data)

    df = pd.DataFrame(data, columns=['Type', 'UID', 'Business_name', 'Company Address', 'District','personal_data', 'Function', 'signature'])

    Convert_to_Excel(df, filename)

def Merge_cells(filename):
    import openpyxl

    # Load the Excel file
    workbook = openpyxl.load_workbook('D:\\Naveed\\Freelancing\\1. Upwork\\2. Work for Portfolio\\2. Jobs to be performed\\Job02\\Program\\Output\\'+filename+'.xlsx')

    # Select the desired worksheet
    worksheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

    # Initialize variables
    current_value = None
    merge_start_row = None
    merge_end_row = None

    # Iterate over the 'Name' column and merge cells with duplicate values
    for row in range(2, worksheet.max_row + 1):
        name_cell = worksheet.cell(row=row, column=2)
        value = name_cell.value

        if value != current_value:
            # Merge cells if there are consecutive duplicate values
            if merge_start_row is not None and merge_end_row is not None:
                worksheet.merge_cells(f'A{merge_start_row}:A{merge_end_row}')

            # Update merge variables for the new set of duplicate values
            current_value = value
            merge_start_row = row
            merge_end_row = row
        else:
            merge_end_row = row

    # Merge cells for the last set of duplicate values, if any
    if merge_start_row is not None and merge_end_row is not None:
        worksheet.merge_cells(f'A{merge_start_row}:A{merge_end_row}')

    # Save the modified workbook
    workbook.save('D:\\Naveed\\Freelancing\\1. Upwork\\2. Work for Portfolio\\2. Jobs to be performed\\Job02\\Program\\Output\\'+filename+'.xlsx')
            
